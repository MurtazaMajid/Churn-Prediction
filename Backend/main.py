"""
Campbell's Restaurant — AI Marketing System
FastAPI Backend

Endpoints:
  GET  /                        → health check
  POST /api/segment             → customer segment + RFM
  POST /api/churn               → churn probability + risk level
  POST /api/sentiment           → ABSA triplets from review text
  POST /api/generate-message    → personalized SMS + email + app notification
  GET  /api/dashboard           → full stats for frontend dashboard
  POST /api/full-pipeline       → all of the above in one call
"""

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List
import pandas as pd
import numpy as np
import pickle
import json
import re
import os
import ast
import requests
import warnings
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.multiclass import OneVsRestClassifier
from sklearn.preprocessing import MultiLabelBinarizer
from xgboost import XGBClassifier
from openpyxl import load_workbook
warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────
# APP SETUP
# ─────────────────────────────────────────────
app = FastAPI(
    title       = "Campbell's AI Marketing API",
    description = "Churn Prediction · Customer Segmentation · ABSA · Personalized Messages",
    version     = "1.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins     = ["*"],  # restrict in production to your Lovable domain
    allow_credentials = True,
    allow_methods     = ["*"],
    allow_headers     = ["*"],
)

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
GROQ_API_KEY  = os.getenv("GROQ_API_KEY", "YOUR_GROQ_API_KEY_HERE")
GROQ_MODEL    = "llama-3.3-70b-versatile"
DATA_DIR      = os.getenv("DATA_DIR", ".")   # folder where CSVs + pickles live
DISCOUNT_MAP  = {'High': 20, 'Medium': 15, 'Low': 10}

# ─────────────────────────────────────────────
# PYDANTIC SCHEMAS
# ─────────────────────────────────────────────
class CustomerRequest(BaseModel):
    recency        : float
    frequency      : int
    monetary       : float
    unique_items   : Optional[int]   = 1
    avg_order_val  : Optional[float] = None
    avg_tip        : Optional[float] = 0.0
    discount_used  : Optional[int]   = 0
    visits_nov     : Optional[int]   = 0
    visits_dec     : Optional[int]   = 0
    visits_jan     : Optional[int]   = 0
    days_since_first: Optional[int]  = None
    favorite_items : Optional[List[str]] = []

class SentimentRequest(BaseModel):
    review: str

class MessageRequest(BaseModel):
    customer_id    : Optional[str]  = None
    segment        : str
    recency        : float
    frequency      : int
    monetary       : float
    risk_level     : str
    churn_probability: float
    favorite_items : Optional[List[str]] = []
    aspects        : Optional[List[str]] = []
    sentiments     : Optional[List[str]] = []

class FullPipelineRequest(BaseModel):
    recency        : float
    frequency      : int
    monetary       : float
    unique_items   : Optional[int]   = 1
    avg_order_val  : Optional[float] = None
    avg_tip        : Optional[float] = 0.0
    discount_used  : Optional[int]   = 0
    visits_nov     : Optional[int]   = 0
    visits_dec     : Optional[int]   = 0
    visits_jan     : Optional[int]   = 0
    days_since_first: Optional[int]  = None
    favorite_items : Optional[List[str]] = []
    review         : Optional[str]   = None

# ─────────────────────────────────────────────
# MODEL STORE — loaded once at startup
# ─────────────────────────────────────────────
models = {}

def load_models():
    """Load all pickle files once at startup."""
    global models
    try:
        with open(f"{DATA_DIR}/kmeans_model.pkl",      'rb') as f: models['kmeans']      = pickle.load(f)
        with open(f"{DATA_DIR}/scaler.pkl",            'rb') as f: models['scaler']      = pickle.load(f)
        with open(f"{DATA_DIR}/cluster_map.pkl",       'rb') as f: models['cluster_map'] = pickle.load(f)
        with open(f"{DATA_DIR}/churn_model_tier2.pkl", 'rb') as f: models['churn']       = pickle.load(f)
        with open(f"{DATA_DIR}/churn_features.pkl",    'rb') as f: models['churn_feats'] = pickle.load(f)
        with open(f"{DATA_DIR}/aspect_model.pkl",      'rb') as f: models['aspect']      = pickle.load(f)
        with open(f"{DATA_DIR}/tfidf_aspect.pkl",      'rb') as f: models['tfidf_asp']   = pickle.load(f)
        with open(f"{DATA_DIR}/mlb.pkl",               'rb') as f: models['mlb']         = pickle.load(f)
        with open(f"{DATA_DIR}/sentiment_model.pkl",   'rb') as f: models['sentiment']   = pickle.load(f)
        with open(f"{DATA_DIR}/tfidf_sent.pkl",        'rb') as f: models['tfidf_sent']  = pickle.load(f)
        print("✅ All models loaded successfully")
    except FileNotFoundError as e:
        print(f"⚠️  Model file not found: {e}. Run the notebook first to generate pickle files.")

def load_menu():
    """Load Campbell's menu data."""
    try:
        wb      = load_workbook(f"{DATA_DIR}/Campbell_Menu_Data_-_2.xlsx", read_only=True)
        ws      = wb.active
        rows    = list(ws.iter_rows(values_only=True))
        menu_df = pd.DataFrame(rows[1:], columns=rows[0])
        food_cats = ['Signature Flights','Brunch Food','Entrées','Desserts',
                     'Salads','Burgers & Sandwiches','Kids Menu',
                     'Sides Dinner','Sides Brunch','Weekly Specials']
        return menu_df[menu_df['Category'].isin(food_cats)].dropna(subset=['itemName','itemPrice'])
    except:
        return pd.DataFrame(columns=['itemName','itemPrice','Category'])

menu_df = pd.DataFrame()

@app.on_event("startup")
async def startup_event():
    global menu_df
    load_models()
    menu_df = load_menu()
    print(f"✅ Menu loaded: {len(menu_df)} items")

# ─────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────
def get_segment(recency, frequency, monetary) -> str:
    """Predict customer segment using loaded KMeans model."""
    if 'kmeans' not in models:
        raise HTTPException(status_code=503, detail="Segmentation model not loaded")
    scaled   = models['scaler'].transform([[recency, frequency, monetary]])
    cluster  = models['kmeans'].predict(scaled)[0]
    return models['cluster_map'].get(cluster, 'Unknown')

def get_churn_score(req: CustomerRequest, segment: str) -> dict:
    """Predict churn probability using two-tier approach."""
    # Tier 1 — rule based (one-time visitors)
    if req.frequency == 1:
        if req.recency < 14:   prob = 0.30
        elif req.recency < 30: prob = 0.65
        else:                  prob = 0.92
        tier = "Rule-Based"
    else:
        # Tier 2 — XGBoost
        if 'churn' not in models:
            raise HTTPException(status_code=503, detail="Churn model not loaded")
        seg_map = {'Regular': 0, 'New': 1, 'Occasional': 2, 'Lost': 3}
        features = {
            'Recency'        : req.recency,
            'Frequency'      : req.frequency,
            'Monetary'       : req.monetary,
            'Unique_Items'   : req.unique_items or 1,
            'Avg_Order_Val'  : req.avg_order_val or req.monetary,
            'Avg_Tip'        : req.avg_tip or 0.0,
            'Discount_Used'  : req.discount_used or 0,
            'Visits_Nov'     : req.visits_nov or 0,
            'Visits_Dec'     : req.visits_dec or 0,
            'Visits_Jan'     : req.visits_jan or 0,
            'Days_Since_First': req.days_since_first or int(req.recency),
            'Segment_Code'   : seg_map.get(segment, 1)
        }
        X     = pd.DataFrame([features])[models['churn_feats']]
        prob  = float(models['churn'].predict_proba(X)[0][1])
        tier  = "XGBoost"

    if prob < 0.33:   risk = "Low"
    elif prob < 0.66: risk = "Medium"
    else:             risk = "High"

    return {"churn_probability": round(prob, 4), "risk_level": risk, "tier": tier}

OPINION_KEYWORDS = {
    'food'    : ['food','dish','meal','taste','flavor','delicious','bland','overcooked','fresh','portion'],
    'staff'   : ['waiter','waitress','server','staff','host','bartender','friendly','rude','helpful','attentive'],
    'service' : ['service','wait','slow','fast','quick','prompt','attentive','responsive'],
    'place'   : ['place','location','restaurant','spot','venue','seating','atmosphere'],
    'menu'    : ['menu','options','variety','selection','choice','specials'],
    'ambience': ['ambience','ambiance','atmosphere','decor','noise','cozy','loud','vibe','setting'],
    'price'   : ['price','expensive','cheap','value','worth','overpriced','affordable','cost']
}

def extract_opinion(review: str, aspect: str) -> str:
    keywords  = OPINION_KEYWORDS.get(aspect, [aspect])
    sentences = re.split(r'[.!?,;]', review)
    for sent in sentences:
        if any(kw in sent.lower() for kw in keywords):
            words = sent.strip().split()
            for i, word in enumerate(words):
                if word.lower().strip('.,!?') in keywords:
                    snippet = ' '.join(words[max(0,i-3):min(len(words),i+4)]).strip('.,!? ')
                    if snippet: return snippet
    return aspect

def run_absa(review: str) -> dict:
    """Run full ABSA pipeline on a review."""
    if 'aspect' not in models:
        raise HTTPException(status_code=503, detail="ABSA model not loaded")
    X_r     = models['tfidf_asp'].transform([review])
    aspects = list(models['mlb'].inverse_transform(models['aspect'].predict(X_r))[0])

    sentiments, opinions, triplets = [], [], []
    for asp in aspects:
        inp       = models['tfidf_sent'].transform([review + ' [ASPECT] ' + asp])
        sentiment = models['sentiment'].predict(inp)[0]
        opinion   = extract_opinion(review, asp)
        sentiments.append(sentiment)
        opinions.append(opinion)
        triplets.append({"aspect": asp, "opinion": opinion, "sentiment": sentiment})

    return {"aspects": aspects, "sentiments": sentiments, "opinions": opinions, "triplets": triplets}

def call_groq(prompt: str) -> str:
    """Call Groq API with JSON mode."""
    headers = {
        "Content-Type" : "application/json",
        "Authorization": f"Bearer {GROQ_API_KEY}"
    }
    body = {
        "model"          : GROQ_MODEL,
        "max_tokens"     : 1000,
        "temperature"    : 0.7,
        "response_format": {"type": "json_object"},
        "messages"       : [
            {
                "role"   : "system",
                "content": "You are a marketing assistant for Campbell's Restaurant. Respond with valid JSON only."
            },
            {"role": "user", "content": prompt}
        ]
    }
    resp = requests.post(
        "https://api.groq.com/openai/v1/chat/completions",
        headers=headers, json=body, timeout=30
    )
    resp.raise_for_status()
    return resp.json()['choices'][0]['message']['content']

def generate_message(req: MessageRequest) -> dict:
    """Build prompt and call Groq to generate personalized messages."""
    discount   = DISCOUNT_MAP.get(req.risk_level, 15)
    liked      = [a for a, s in zip(req.aspects or [], req.sentiments or []) if s == 'positive']
    disliked   = [a for a, s in zip(req.aspects or [], req.sentiments or []) if s == 'negative']
    highlights = []
    if not menu_df.empty:
        highlights = menu_df.sample(min(3, len(menu_df)))[['itemName','itemPrice','Category']].to_dict('records')

    prompt = f"""You are a warm, friendly marketing assistant for Campbell's Restaurant.

CUSTOMER PROFILE:
- Segment            : {req.segment}
- Days since last visit: {int(req.recency)} days
- Total visits       : {req.frequency}
- Total spent        : ${req.monetary:.2f}
- Churn risk         : {req.risk_level}
- Liked aspects      : {liked if liked else 'unknown'}
- Disliked aspects   : {disliked if disliked else 'none noted'}
- Favorite items     : {req.favorite_items}

MENU HIGHLIGHTS:
{json.dumps(highlights, indent=2)}

DISCOUNT TO OFFER: {discount}% off next visit

Return ONLY this JSON:
{{
  "sms": "...",
  "email": {{"subject": "...", "body": "..."}},
  "app_notification": "..."
}}

RULES:
- SMS: max 160 chars, casual and punchy
- Email: warm, personal, 3-4 short paragraphs
- App notification: max 80 chars, exciting
- Tone: Lost→urgent ("we miss you!"), New→welcoming, Occasional→appreciative, Regular→VIP
- NEVER mention churn, AI, or risk scores"""

    response = call_groq(prompt)
    try:
        clean      = re.sub(r'```json|```', '', response).strip()
        json_match = re.search(r'\{.*\}', clean, re.DOTALL)
        if json_match: clean = json_match.group(0)
        return json.loads(clean)
    except:
        return {"sms": response[:160], "email": {}, "app_notification": response[:80]}

# ─────────────────────────────────────────────
# ENDPOINTS
# ─────────────────────────────────────────────

@app.get("/")
def health_check():
    return {
        "status" : "ok",
        "service": "Campbell's AI Marketing API",
        "version": "1.0.0",
        "models_loaded": list(models.keys())
    }

# ── Segmentation ──
@app.post("/api/segment")
def segment_customer(req: CustomerRequest):
    """
    Segment a customer into New / Occasional / Regular / Lost
    based on their RFM (Recency, Frequency, Monetary) values.
    """
    segment = get_segment(req.recency, req.frequency, req.monetary)
    return {
        "segment" : segment,
        "recency" : req.recency,
        "frequency": req.frequency,
        "monetary": req.monetary
    }

# ── Churn Prediction ──
@app.post("/api/churn")
def predict_churn(req: CustomerRequest):
    """
    Predict churn probability for a customer.
    Uses rule-based scoring for one-time visitors,
    XGBoost for repeat customers.
    """
    segment = get_segment(req.recency, req.frequency, req.monetary)
    result  = get_churn_score(req, segment)
    return {
        "segment"          : segment,
        "churn_probability": result['churn_probability'],
        "risk_level"       : result['risk_level'],
        "tier"             : result['tier'],
        "discount_to_offer": f"{DISCOUNT_MAP.get(result['risk_level'], 15)}%"
    }

# ── Sentiment Analysis ──
@app.post("/api/sentiment")
def analyze_sentiment(req: SentimentRequest):
    """
    Run Aspect-Based Sentiment Analysis on a restaurant review.
    Returns detected aspects, sentiment per aspect, opinion phrases, and triplets.
    """
    result = run_absa(req.review)
    return {
        "review"    : req.review,
        "aspects"   : result['aspects'],
        "sentiments": result['sentiments'],
        "opinions"  : result['opinions'],
        "triplets"  : result['triplets']
    }

# ── Message Generation ──
@app.post("/api/generate-message")
def generate_personalized_message(req: MessageRequest):
    """
    Generate personalized re-engagement messages (SMS, email, app notification)
    for a customer using Groq LLaMA 3.3 70B.
    """
    messages = generate_message(req)
    return {
        "customer_id"     : req.customer_id,
        "segment"         : req.segment,
        "risk_level"      : req.risk_level,
        "discount_offered": f"{DISCOUNT_MAP.get(req.risk_level, 15)}%",
        "messages"        : messages
    }

# ── Full Pipeline ──
@app.post("/api/full-pipeline")
def full_pipeline(req: FullPipelineRequest):
    """
    Run the complete pipeline for a single customer in one call:
    1. Segment → 2. Churn Score → 3. ABSA (if review provided) → 4. Generate Message
    Lovable frontend calls this single endpoint for the full experience.
    """
    # Step 1 — Segment
    segment = get_segment(req.recency, req.frequency, req.monetary)

    # Step 2 — Churn
    churn_req = CustomerRequest(**req.dict())
    churn     = get_churn_score(churn_req, segment)

    # Step 3 — ABSA (optional)
    absa = None
    if req.review:
        absa = run_absa(req.review)

    # Step 4 — Generate message
    msg_req = MessageRequest(
        segment          = segment,
        recency          = req.recency,
        frequency        = req.frequency,
        monetary         = req.monetary,
        risk_level       = churn['risk_level'],
        churn_probability= churn['churn_probability'],
        favorite_items   = req.favorite_items,
        aspects          = absa['aspects']    if absa else [],
        sentiments       = absa['sentiments'] if absa else []
    )
    messages = generate_message(msg_req)

    return {
        "segment"          : segment,
        "churn_probability": churn['churn_probability'],
        "risk_level"       : churn['risk_level'],
        "tier"             : churn['tier'],
        "discount_offered" : f"{DISCOUNT_MAP.get(churn['risk_level'], 15)}%",
        "absa"             : absa,
        "messages"         : messages
    }

# ── Dashboard Stats ──
@app.get("/api/dashboard")
def get_dashboard_stats():
    """
    Return aggregated stats for the Lovable dashboard.
    Loads churn_scores_final.csv generated by the notebook.
    """
    try:
        df = pd.read_csv(f"{DATA_DIR}/churn_scores_final.csv")
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="churn_scores_final.csv not found. Run notebook first.")

    # Segment counts
    seg_counts = df['Segment'].value_counts().to_dict()

    # Risk breakdown
    risk_counts = df['Risk_Level'].astype(str).value_counts().to_dict()

    # Churn rate
    churn_rate = round(df['Churn'].mean() * 100, 2) if 'Churn' in df.columns else None

    # Top at-risk customers
    top_risk = (
        df[df['Risk_Level'].astype(str) == 'High']
        .sort_values('Churn_Probability', ascending=False)
        .head(20)[['Last 4 Card Digits','Segment','Recency','Frequency','Monetary','Churn_Probability','Risk_Level','Tier']]
        .to_dict('records')
    )

    # Avg RFM per segment
    rfm_by_segment = (
        df.groupby('Segment')[['Recency','Frequency','Monetary']]
        .mean().round(2)
        .to_dict('index')
    )

    return {
        "total_customers"  : int(len(df)),
        "segment_counts"   : seg_counts,
        "risk_counts"      : risk_counts,
        "churn_rate_pct"   : churn_rate,
        "top_at_risk"      : top_risk,
        "rfm_by_segment"   : rfm_by_segment,
    }
